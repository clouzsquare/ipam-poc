from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from io import BytesIO
from typing import Dict, List, Union
import json
import operator
import os
import re
import smtplib

from dotenv import load_dotenv
from langchain_google_genai import ChatGoogleGenerativeAI
from langchain_core.messages import AIMessage, BaseMessage, HumanMessage, SystemMessage
from langgraph.graph import END, StateGraph
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.client.ntoss_client import NtossClient
from app.core.database import SessionLocal
from app.repositories.candidate.candidate_repository import CandidateRepository
from .shared_state import AgentState


class CandidateAgent:
    REQUIRED_HEADERS = [
        "DHCP Server IP",
        "IP블록",
        "인프라팀",
        "네트워크 이름",
        "네트워크 ID",
        "Primary 여부",
        "사용률(%)",
    ]

    def __init__(self):
        load_dotenv()
        self.ntoss = NtossClient()
        self.llm = ChatGoogleGenerativeAI(
            model="gemini-2.5-flash",
            temperature=0,
            google_api_key=os.getenv("GOOGLE_API_KEY"),
        )
        self._classification_cache: Dict[str, bool] = {}

    def _convert_to_messages(self, messages: List[Union[dict, BaseMessage]]) -> List[BaseMessage]:
        converted: List[BaseMessage] = []
        for m in messages:
            if isinstance(m, BaseMessage):
                converted.append(m)
                continue
            role = m.get("role", "user")
            content = m.get("content", "")
            converted.append(HumanMessage(content=content) if role == "user" else AIMessage(content=content))
        return converted

    def infer_upload_mode_from_history(self, history: List[dict]) -> str:
        """채팅 컨텍스트를 보고 업로드 목적(extract/finalize)을 판별합니다."""
        converted = self._convert_to_messages(history or [])
        if not converted:
            return "extract"

        prompt = """
        당신은 엑셀파일 업로드 목적 판별기입니다.
        최근 대화 맥락을 바탕으로 아래 중 무엇인지 판단하세요.
        - EXTRACT: "추출", "목록 추출", "후보 추출" 등 IP회수 후보 목록 추출/미리보기 목적의 엑셀 업로드
        - FINALIZE: "확정", "최종", "반영" 등 IP회수 후보 목록 확정 및 DB 반영 목적의 엑셀 업로드

        출력은 EXTRACT 또는 FINALIZE 한 단어만 반환하세요.
        """
        try:
            res = self.llm.invoke([SystemMessage(content=prompt)] + converted[-6:])
            text = str(res.content).upper()
            if "FINALIZE" in text:
                return "finalize"
            return "extract"
        except Exception:
            # LLM 실패 시 보수적 기본값
            last_user = ""
            for m in reversed(history or []):
                if m.get("role") == "user":
                    last_user = str(m.get("content", ""))
                    break
            if any(k in last_user for k in ["확정", "최종", "DB 반영"]):
                return "finalize"
            return "extract"

    @staticmethod
    def _extract_json(raw_text: str, fallback: Dict) -> Dict:
        try:
            cleaned = re.sub(r"```json|```", "", raw_text).strip()
            return json.loads(cleaned)
        except Exception:
            return fallback

    def intent_analyzer(self, state: AgentState):
        """[1] 의도 분석"""
        print("\n🚀 [NODE: intent_analyzer(Candidate)]")
        history = self._convert_to_messages(state["messages"])
        system_prompt = """
        당신은 IP 회수 후보 추출 에이전트 의도 분류기입니다.
        - START: "후보 목록 추출", "후보 뽑아줘", "추출해줘" 등 시점 기준으로 후보 추출을 새롭게 시작할 때
        - PROVIDE: "업로드 완료", "파일 넣었어", "내용 입력했어" 등 엑셀 파일을 업로드했을 때
        - CONFIRM: "예", "진행", "검토 진행", "메일 발송" 등 인프라담당자 검토 요청 메일을 발송할 때
        - FINALIZE: "확정", "IP회수 후보 확정", "후보 확정", "최종 확정" 등 회수 후보를 확정할 때
        - REJECT: "아니오", "취소", "보류", "다시" 등 회수 후보를 다시 추출할 때
        - STATUS: "현황", "목록", "조회" 등 현재 후보 목록을 조회할 때
        - CHAT: 그 외
        반드시 START, PROVIDE, CONFIRM, FINALIZE, REJECT, STATUS, CHAT 중 하나만 답변하세요.
        """
        res = self.llm.invoke([SystemMessage(content=system_prompt)] + history)
        raw = str(res.content).upper()
        intent = next((x for x in ["START", "PROVIDE", "CONFIRM", "FINALIZE", "REJECT", "STATUS"] if x in raw), "CHAT")
        print(f"🎯 분석된 Intent: {intent}")
        return {"intent": intent}

    def action_planner(self, state: AgentState):
        """[2] 액션 계획 수립: intent를 실행 액션으로 변환"""
        print("\n🚀 [NODE: action_planner(Candidate)]")
        intent = state.get("intent", "CHAT")
        history = self._convert_to_messages(state["messages"])
        prompt = f"""
        당신은 후보 추출 정책 설계자입니다. intent={intent}에 맞춰 JSON만 출력하세요.

        규칙:
        - START: 업로드 요청 안내
        - PROVIDE: 최신 후보 미리보기 조회 및 검토 요청
        - CONFIRM: 메일 발송 실행
        - FINALIZE: 확정용 엑셀 업로드 안내
        - REJECT: 추가 후보 요청 유도
        - STATUS: 최근 후보 목록 조회
        - CHAT: 빈 계획.

        JSON 스키마:
        {{
          "action": "GUIDE_UPLOAD|FETCH_CANDIDATES|SEND_REVIEW_MAIL|GUIDE_FINALIZE_UPLOAD|ASK_MORE_TARGET|FETCH_STATUS|CHAT"
        }}
        """
        res = self.llm.invoke([SystemMessage(content=prompt)] + history)
        default_plan = {"action": "CHAT"}
        plan = self._extract_json(str(res.content), default_plan)
        action = str(plan.get("action", "CHAT")).upper()
        intent_to_action = {
            "START": "GUIDE_UPLOAD",
            "PROVIDE": "FETCH_CANDIDATES",
            "CONFIRM": "SEND_REVIEW_MAIL",
            "FINALIZE": "GUIDE_FINALIZE_UPLOAD",
            "REJECT": "ASK_MORE_TARGET",
            "STATUS": "FETCH_STATUS",
            "CHAT": "CHAT",
        }
        if action not in {"GUIDE_UPLOAD", "FETCH_CANDIDATES", "SEND_REVIEW_MAIL", "GUIDE_FINALIZE_UPLOAD", "ASK_MORE_TARGET", "FETCH_STATUS", "CHAT"}:
            action = intent_to_action.get(intent, "CHAT")
        print(f"🎯 분석된 action: {action}")
        return {"query_plan": {"action": action}}

    def data_fetcher(self, state: AgentState):
        """[3] 데이터 조회: 액션에 필요한 후보 데이터만 수집"""
        print("\n🚀 [NODE: data_fetcher(Candidate)]")
        plan = state.get("query_plan", {})
        action = plan.get("action")
        if action not in {"FETCH_STATUS", "FETCH_CANDIDATES"}:
            return {"selected_candidates": [], "selected_ips": []}

        if action == "FETCH_CANDIDATES":
            selected = state.get("selected_ips", [])
            return {"selected_candidates": selected, "selected_ips": selected}

        db = SessionLocal()
        try:
            repo = CandidateRepository(db)
            selected_candidates = repo.get_all_candidates_latest()
            return {
                "selected_candidates": selected_candidates,
                "selected_ips": selected_candidates,
            }
        finally:
            db.close()

    @staticmethod
    def _load_team_email_map() -> Dict[str, str]:
        raw = os.getenv("INFRA_TEAM_EMAIL_MAP", "").strip()
        if not raw:
            return {}
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, dict):
                return {str(k).strip(): str(v).strip() for k, v in parsed.items() if str(v).strip()}
        except Exception:
            pass
        mapping: Dict[str, str] = {}
        for pair in raw.split(";"):
            if ":" not in pair:
                continue
            key, value = pair.split(":", 1)
            if key.strip() and value.strip():
                mapping[key.strip()] = value.strip()
        return mapping

    @staticmethod
    def _build_review_excel_bytes(selected_ips: List[dict]) -> bytes:
        """업로드 원본 엑셀과 동일한 컬럼(행 단위 스냅샷)이 있으면 그대로, 없으면 최소 컬럼으로 .xlsx 생성."""
        wb = Workbook()
        ws = wb.active
        with_snap = [x for x in selected_ips if x.get("excel_row")]
        if with_snap:
            hdrs = list(with_snap[0]["excel_row"].keys())
            ws.append(hdrs)
            for item in with_snap:
                er = item["excel_row"]
                ws.append([er.get(h) for h in hdrs])
        else:
            ws.append(["네트워크 ID", "IP블록", "인프라팀", "담당 이메일"])
            for item in selected_ips:
                ws.append(
                    [
                        item.get("nw_id"),
                        item.get("ip_address"),
                        item.get("owner_team"),
                        item.get("owner_email"),
                    ]
                )
        bio = BytesIO()
        wb.save(bio)
        return bio.getvalue()

    def _send_review_mails(self, selected_ips: List[dict]) -> Dict:
        if not selected_ips:
            return {"sent_count": 0, "failed": []}

        team_email_map = self._load_team_email_map()
        default_email = os.getenv("CANDIDATE_DEFAULT_OWNER_EMAIL", "no-reply@ipam.local")
        recipients = sorted(
            {
                team_email_map.get(item.get("owner_team", "").strip()) or item.get("owner_email") or default_email
                for item in selected_ips
                if item.get("owner_team") or item.get("owner_email")
            }
        )
        gmail_user = os.getenv("GMAIL_USER")
        gmail_password = os.getenv("GMAIL_APP_PASSWORD")
        subject = "[IPAM] IP 회수 후보 검토 요청"

        body_lines = [
            "안녕하세요. IPAM AI Agent입니다.",
            "아래 IP 회수 후보에 대한 검토를 요청드립니다.",
            "첨부 엑셀은 선정된 회수 후보만 포함합니다(제외 행 제거).",
            "",
        ]
        for item in selected_ips[:30]:
            body_lines.append(f"- {item.get('owner_team')} | {item.get('nw_id')} | {item.get('ip_address')}")
        body_lines.append("")
        body_lines.append("검토 후 회신 부탁드립니다.")
        body = "\n".join(body_lines)

        if not gmail_user or not gmail_password:
            # PoC 모드: 실제 SMTP 미설정 시 mock 성공 처리
            return {"sent_count": len(recipients), "failed": []}

        xlsx_bytes = self._build_review_excel_bytes(selected_ips)
        attach_name = "ip_reclaim_candidates_review.xlsx"

        failed = []
        for to_email in recipients:
            msg = MIMEMultipart()
            msg["Subject"] = subject
            msg["From"] = gmail_user
            msg["To"] = to_email
            msg.attach(MIMEText(body, _charset="utf-8"))
            part = MIMEApplication(
                xlsx_bytes,
                _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            part.add_header("Content-Disposition", "attachment", filename=attach_name)
            msg.attach(part)
            try:
                with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
                    smtp.login(gmail_user, gmail_password)
                    smtp.send_message(msg)
            except Exception:
                failed.append(to_email)
        return {"sent_count": len(recipients) - len(failed), "failed": failed}

    def _insert_confirmed_candidates(self, db: Session, selected_ips: List[dict], extraction_batch_id: str = "") -> Dict:
        if not selected_ips:
            return {"inserted_count": 0, "skipped_count": 0}
        normalized = []
        fallback_email = os.getenv("CANDIDATE_DEFAULT_OWNER_EMAIL", "no-reply@ipam.local")
        for item in selected_ips:
            copied = dict(item)
            copied["owner_email"] = str(copied.get("owner_email", "")).strip() or fallback_email
            normalized.append(copied)
        repo = CandidateRepository(db)
        return repo.insert_confirmed_candidates(normalized, extraction_batch_id)

    def responder(self, state: AgentState):
        print("🚀 [NODE: responder(Candidate)]")
        action = state.get("query_plan", {}).get("action", "CHAT")
        if action == "GUIDE_UPLOAD":
            msg = (
                "IP회수 후보 목록 추출을 시작합니다. 먼저 NW ID별 IP대역 사용률 엑셀파일이 필요합니다.\n"
                "엑셀 업로드 API로 파일을 전달해 주세요.\n"
                "- POST /api/v1/candidate/extract\n"
                "- 필수: file(.xlsx)\n"
                "- 선택: usage_threshold, extraction_batch_id, default_owner_email\n"
                "업로드가 완료되면 후보 목록을 바로 확인할 수 있습니다."
            )
            return {"messages": [AIMessage(content=msg)]}

        if action in {"FETCH_STATUS", "FETCH_CANDIDATES"}:
            data = state.get("selected_candidates", []) or state.get("selected_ips", [])
            if not data:
                return {"messages": [AIMessage(content="업로드된 후보 목록이 없습니다. 엑셀 파일을 업로드해 주세요.")]}
            summarize_prompt = (
                "당신은 IPAM AI Assistant입니다.\n"
                "아래 후보 목록을 간단 요약하고, 마지막 줄에 "
                "'후보 확인 후 \"메일 발송\"이라고 입력하면 검토 메일을 인프라 담당자에게 발송하고, "
                "수정이 필요하다면 수정할 내용을 입력해주세요'를 반드시 포함하세요.\n"
                f"데이터: {data}"
            )
            res = self.llm.invoke([HumanMessage(content=summarize_prompt)])
            return {"messages": [AIMessage(content=str(res.content))], "selected_ips": data}

        if action == "SEND_REVIEW_MAIL":
            selected_ips = state.get("selected_ips", [])
            if not selected_ips:
                return {"messages": [AIMessage(content="검토 메일을 보낼 후보 목록이 없습니다. 먼저 엑셀 업로드를 진행해 주세요.")]}
            mail_result = self._send_review_mails(selected_ips)
            if mail_result["failed"]:
                return {"messages": [AIMessage(content=f"일부 메일 발송 실패: {mail_result['failed']}")]}
            return {"messages": [AIMessage(content=f"인프라 담당자 검토 요청 메일 발송 완료 ({mail_result['sent_count']}명)")]}

        if action == "GUIDE_FINALIZE_UPLOAD":
            return {
                "messages": [
                    AIMessage(
                        content=(
                            "IP회수 후보 목록 확정을 진행합니다. 인프라 담당자 검토가 반영된 엑셀파일을 업로드해 주세요.\n"
                            "- POST /api/v1/candidate/finalize\n"
                            "- 해당 업로드는 즉시 DB에 INSERT 됩니다."
                        )
                    )
                ]
            }

        if action == "ASK_MORE_TARGET":
            return {
                "messages": [
                    AIMessage(
                        content=(
                            "확정하지 않았습니다. 추가로 회수할 대상을 지정해 주세요.\n"
                            "예: 특정 팀, 특정 NW ID, 사용률 임계치 조정 요청"
                        )
                    )
                ]
            }

        return {"messages": [AIMessage(content="후보 추출 관련 문의를 도와드릴게요.")] }

    @staticmethod
    def _normalize_header(value) -> str:
        return str(value).strip() if value is not None else ""

    @staticmethod
    def _to_percent(value) -> float:
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value) * 100 if float(value) <= 1 else float(value)
        text = str(value).strip().replace("%", "")
        try:
            return float(text)
        except ValueError:
            return 0.0

    @staticmethod
    def _is_non_primary(value) -> bool:
        return str(value).strip().upper() != "Y" if value is not None else True

    def _is_accommodation_by_llm(self, name: str) -> bool:
        normalized = (name or "").strip()
        if not normalized:
            return False
        if normalized in self._classification_cache:
            return self._classification_cache[normalized]

        prompt = f"""
        아래 명칭이 '특정 요일/시점에 사용량이 급증·급감할 수 있는 단기 숙박 시설'인지 분류하세요.
        단기 숙박 시설 예: 기숙사, 호텔, 숙박업소, 모텔, 리조트, 게스트하우스 등
        일반 주거 아파트/일반 상업시설은 KEEP으로 분류하세요.
        명칭: "{normalized}"
        출력은 EXCLUDE 또는 KEEP 한 단어만 반환하세요.
        """
        try:
            response = self.llm.invoke(prompt)
            is_excluded = "EXCLUDE" in str(response.content).upper()
        except Exception:
            is_excluded = False
        self._classification_cache[normalized] = is_excluded
        return is_excluded

    def _llm_generate_reason(self, row_context: Dict, excluded: bool) -> str:
        prompt = f"""
        아래 판정 결과를 관리자에게 설명할 한 줄 사유를 작성하세요.
        - excluded={excluded}
        - 판정 기준: 사용률 임계치, Primary 여부, 숙소형 시설 제외
        - 데이터: {row_context}
        40자 이내 한국어 문장으로 출력하세요.
        """
        try:
            res = self.llm.invoke(prompt)
            return str(res.content).strip()
        except Exception:
            return "정책 기준에 따라 자동 판정됨"

    def extract_candidates_from_excel(
        self,
        db: Session,
        file_bytes: bytes,
        extraction_batch_id: str,
        usage_threshold: float,
        default_owner_email: str,
    ) -> Dict:
        wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError("엑셀 파일이 비어 있습니다.")

        headers = [self._normalize_header(h) for h in rows[0]]
        header_index = {name: idx for idx, name in enumerate(headers)}
        missing_headers = [h for h in self.REQUIRED_HEADERS if h not in header_index]
        if missing_headers:
            raise ValueError(f"필수 컬럼이 없습니다: {', '.join(missing_headers)}")

        inserted = 0
        skipped = 0
        excluded_by_accommodation = 0
        selected_preview: List[Dict] = []
        excluded_details: List[Dict] = []
        selected_ips: List[Dict] = []
        seen_keys: set[Tuple[str, str]] = set()

        for row in rows[1:]:
            dhcp_ip = row[header_index["DHCP Server IP"]]
            ip_block = str(row[header_index["IP블록"]] or "").strip()
            owner_team = row[header_index["인프라팀"]]
            network_name = str(row[header_index["네트워크 이름"]] or "")
            nw_id = str(row[header_index["네트워크 ID"]] or "")
            primary_flag = row[header_index["Primary 여부"]]
            usage_raw = row[header_index["사용률(%)"]]
            display_ip = ip_block if ip_block else f"{dhcp_ip}/32"

            if not dhcp_ip or not nw_id or not owner_team:
                skipped += 1
                continue

            usage_percent = self._to_percent(usage_raw)
            ntoss_result = self.ntoss.get_apartment_info_by_nw_id(nw_id=nw_id)
            apartment_name = str(ntoss_result.get("apartment_name", "") or "")
            network_is_accommodation = self._is_accommodation_by_llm(network_name)
            apartment_is_accommodation = self._is_accommodation_by_llm(apartment_name)

            row_context = {
                "nw_id": nw_id,
                "ip_address": display_ip,
                "owner_team": str(owner_team),
                "usage_percent": usage_percent,
                "is_non_primary": self._is_non_primary(primary_flag),
                "network_name": network_name,
                "apartment_name": apartment_name,
                "network_name_is_accommodation": network_is_accommodation,
                "apartment_name_is_accommodation": apartment_is_accommodation,
            }
            is_under_threshold = usage_percent < usage_threshold
            is_non_primary = self._is_non_primary(primary_flag)
            is_excluded_accommodation = network_is_accommodation or apartment_is_accommodation
            should_select = is_under_threshold and is_non_primary and (not is_excluded_accommodation)

            if not should_select:
                skipped += 1
                if is_excluded_accommodation:
                    excluded_by_accommodation += 1
                reasons = []
                if not is_under_threshold:
                    reasons.append(f"사용률 {usage_percent:.2f}%가 기준({usage_threshold:.2f}%) 미만이 아님")
                if not is_non_primary:
                    reasons.append("Primary 여부가 Y이므로 제외")
                if is_excluded_accommodation:
                    reasons.append("네트워크명 또는 NTOSS 아파트명이 숙소형 시설로 분류됨")
                excluded_details.append(
                    {
                        "nw_id": nw_id,
                        "ip_address": display_ip,
                        "owner_team": str(owner_team),
                        "usage_percent": usage_percent,
                        "network_name": network_name,
                        "apartment_name": apartment_name,
                        "exclude_reason": " / ".join(reasons) if reasons else "정책 기준 미충족",
                    }
                )
                continue

            unique_key = (nw_id, display_ip)
            if unique_key in seen_keys:
                skipped += 1
                excluded_details.append(
                    {
                        "nw_id": nw_id,
                        "ip_address": display_ip,
                        "owner_team": str(owner_team),
                        "usage_percent": usage_percent,
                        "network_name": network_name,
                        "apartment_name": apartment_name,
                        "exclude_reason": "엑셀 내 중복 대상",
                    }
                )
                continue
            seen_keys.add(unique_key)
            inserted += 1
            excel_row = {
                h: row[header_index[h]] if header_index[h] < len(row) else None for h in headers
            }
            selected_item = {
                "nw_id": nw_id,
                "ip_address": display_ip,
                "owner_team": str(owner_team),
                "owner_email": default_owner_email,
                "usage_percent": usage_percent,
                "network_name": network_name,
                "apartment_name": apartment_name,
                "decision_reason": self._llm_generate_reason(row_context, excluded=False),
            }
            selected_ips.append(
                {
                    "nw_id": selected_item["nw_id"],
                    "ip_address": selected_item["ip_address"],
                    "owner_team": selected_item["owner_team"],
                    "owner_email": selected_item["owner_email"],
                    "excel_row": excel_row,
                }
            )
            selected_preview.append(
                selected_item
            )

        return {
            "batch_id": extraction_batch_id,
            "usage_threshold": usage_threshold,
            "selected_count": inserted,
            "skipped_count": skipped,
            "excluded_by_accommodation_count": excluded_by_accommodation,
            "selection_policy": {
                "usage_threshold_percent": usage_threshold,
                "non_primary_required": True,
                "exclude_accommodation": True,
            },
            "selected_preview": selected_preview,
            "excluded_details": excluded_details,
            "selected_ips": selected_ips,
            "requires_finalize": True,
        }

    def finalize_candidates_from_excel(
        self,
        db: Session,
        file_bytes: bytes,
        extraction_batch_id: str,
        usage_threshold: float,
        default_owner_email: str,
    ) -> Dict:
        extracted = self.extract_candidates_from_excel(
            db=db,
            file_bytes=file_bytes,
            extraction_batch_id=extraction_batch_id,
            usage_threshold=usage_threshold,
            default_owner_email=default_owner_email,
        )
        insert_result = self._insert_confirmed_candidates(
            db=db,
            selected_ips=extracted.get("selected_ips", []),
            extraction_batch_id=extraction_batch_id,
        )
        return {
            "batch_id": extraction_batch_id,
            "selected_count": extracted.get("selected_count", 0),
            "excluded_by_accommodation_count": extracted.get("excluded_by_accommodation_count", 0),
            "inserted_count": insert_result.get("inserted_count", 0),
            "skipped_count": insert_result.get("skipped_count", 0),
        }

    def build_extract_response_message(self, result: Dict) -> str:
        prompt = f"""
        당신은 IPAM AI Assistant입니다.
        아래 데이터를 바탕으로 "정해진 양식"으로만 응답하세요.

        [중요 규칙]
        1) 후보 목록(selected_preview)과 제외 목록(excluded_details)을 절대 요약/생략하지 말고 전부 출력하세요.
        2) 기준 IP사용률(usage_threshold_percent)을 반드시 명시하세요.
        3) 제외 목록은 각 항목의 exclude_reason을 그대로 포함하세요.
        4) 데이터에 없는 내용을 임의로 만들지 마세요.
        5) 아래 출력 템플릿의 제목/순서를 그대로 지키세요.
        6) 마지막 안내 문장은 반드시 아래 문장과 100% 동일해야 합니다.
           후보 확인 후 '메일 발송'이라고 입력하면 검토 메일을 인프라 담당자에게 발송하고, 수정이 필요하다면 수정할 내용을 입력해주세요.

        [출력 템플릿]
        엑셀 분석 결과 요약
        - 후보 건수: {{selected_count}}건
        - 제외 건수: {{skipped_count}}건
        - 기준 IP사용률: {{usage_threshold_percent}}%
        - 선정 기준: 사용률 미달 + Non-primary + 단기 숙박 시설 제외

        후보 목록
        - {{owner_team}} | {{nw_id}} | {{ip_address}} | 사용률 {{usage_percent}}% | 근거: {{decision_reason}}
        - ... (selected_preview의 모든 항목)

        제외 목록
        - {{owner_team}} | {{nw_id}} | {{ip_address}} | 사용률 {{usage_percent}}% | 제외 사유: {{exclude_reason}}
        - ... (excluded_details의 모든 항목)

        후보 확인 후 '메일 발송'이라고 입력하면 검토 메일을 인프라 담당자에게 발송하고, 수정이 필요하다면 수정할 내용을 입력해주세요.

        [입력 데이터]
        {result}
        """
        try:
            res = self.llm.invoke([HumanMessage(content=prompt)])
            return str(res.content).strip()
        except Exception:
            selected_preview = result.get("selected_preview", []) or []
            excluded_details = result.get("excluded_details", []) or []
            usage_threshold = (
                result.get("selection_policy", {}) or {}
            ).get("usage_threshold_percent", "-")

            lines = [
                "엑셀 분석 결과 요약",
                f"- 후보 건수: {result.get('selected_count', 0)}건",
                f"- 제외 건수: {result.get('skipped_count', 0)}건",
                f"- 기준 IP사용률: {usage_threshold}%",
                "- 선정 기준: 사용률 미달 + Non-primary + 단기 숙박 시설 제외",
                "",
                "후보 목록",
            ]

            if selected_preview:
                for item in selected_preview:
                    lines.append(
                        f"- {item.get('owner_team')} | {item.get('nw_id')} | {item.get('ip_address')} | "
                        f"사용률 {item.get('usage_percent')}% | 근거: {item.get('decision_reason', '정책 기준 충족')}"
                    )
            else:
                lines.append("- 후보 없음")

            lines.append("")
            lines.append("제외 목록")
            if excluded_details:
                for item in excluded_details:
                    lines.append(
                        f"- {item.get('owner_team')} | {item.get('nw_id')} | {item.get('ip_address')} | "
                        f"사용률 {item.get('usage_percent')}% | 제외 사유: {item.get('exclude_reason', '정책 기준 미충족')}"
                    )
            else:
                lines.append("- 제외 없음")

            lines.append("")
            lines.append(
                "후보 확인 후 '메일 발송'이라고 입력하면 검토 메일을 인프라 담당자에게 발송하고, 수정이 필요하다면 수정할 내용을 입력해주세요."
            )
            return "\n".join(lines)

    def build_finalize_response_message(self, result: Dict) -> str:
        prompt = f"""
        당신은 IPAM AI Assistant입니다.
        아래 확정 결과를 한 문단으로 사용자에게 보고하세요.
        데이터: {result}
        """
        try:
            res = self.llm.invoke([HumanMessage(content=prompt)])
            return str(res.content).strip()
        except Exception:
            return (
                f"후보 확정 완료: 추출 {result.get('selected_count', 0)}건 중 "
                f"DB INSERT {result.get('inserted_count', 0)}건, 제외 {result.get('skipped_count', 0)}건"
            )


def build_candidate_graph():
    agent = CandidateAgent()
    workflow = StateGraph(AgentState)
    workflow.add_node("analyzer", agent.intent_analyzer)
    workflow.add_node("constructor", agent.action_planner)
    workflow.add_node("fetcher", agent.data_fetcher)
    workflow.add_node("responder", agent.responder)

    workflow.set_entry_point("analyzer")
    workflow.add_edge("analyzer", "constructor")
    workflow.add_conditional_edges(
        "constructor",
        lambda x: x.get("query_plan", {}).get("action", "CHAT"),
        {
            "FETCH_STATUS": "fetcher",
            "FETCH_CANDIDATES": "fetcher",
            "GUIDE_UPLOAD": "responder",
            "SEND_REVIEW_MAIL": "responder",
            "GUIDE_FINALIZE_UPLOAD": "responder",
            "ASK_MORE_TARGET": "responder",
            "CHAT": "responder",
        },
    )
    workflow.add_edge("fetcher", "responder")
    workflow.add_edge("responder", END)
    return workflow.compile()


candidate_graph = build_candidate_graph()